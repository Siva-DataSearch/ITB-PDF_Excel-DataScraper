import PyPDF2
import os
import re
import openpyxl

def get_pdf_files(path):
    pdf_files = []
    for root, dirs, files in os.walk(path):
        for file in files:
            if file.endswith(".pdf"):
                pdf_files.append(os.path.join(root, file))
    return pdf_files


Folder_name = "PDF_Files"
pdf_files_list = get_pdf_files(Folder_name)
Results = []
keywords = ['phone', 'web', 'email', 'linkedin', 'fax', 'facebook', 'youtube', 'instagram', 'Twitter']
column_names = ['Title', 'country', 'email', 'Twitter', 'web', 'phone', 'instagram', 'linkedin', 'facebook',
                'fax', 'Address', 'youtube']


def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ''
        for page in reader.pages:
            text += page.extract_text()
        return text


for index, file in enumerate(pdf_files_list):
    text = extract_text_from_pdf(file).replace('\x0c', '').strip()
    # print(text)

    original_list = list(text.split("\n"))
    stripped_list = [value.strip() for value in original_list]
    # print(stripped_list)

    # for index, value in enumerate(stripped_list, start=1):
    #     print(index, value)

    Title = stripped_list[1]
    # Address = stripped_list[2:5]
    # print("Title-", str(Title[0]))
    # print("Title-", Title)

    found_index = float('inf')
    found_keyword = None

    for keyword in keywords:
        for index, element in enumerate(stripped_list):
            if keyword.lower() in element.lower():
                if index < found_index:
                    found_index = index
                    found_keyword = keyword

    if found_keyword:
        # print(f"The first occurrence of '{found_keyword}' is at index: {found_index}")
        Address = stripped_list[2:found_index]
        country = Address[-1]
        # print("Country:", country)

    else:
        # print("No keywords found in the list.")
        Address = stripped_list[2:5]
        country = Address[-1]

    result = {}
    result['Title'] = Title
    result['Address'] = ','.join(Address)
    result['country'] = country

    for item in stripped_list:
        for keyword in keywords:
            pattern = r'\b{}:\s*(.*)'.format(keyword)
            match = re.search(pattern, item, re.IGNORECASE)
            if match:
                value = match.group(1)
                result[keyword] = value

    # print(result)
    Results.append(result)

print(len(Results))

# Create a new workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Results'

# Writing column names to the first row
for col, column_name in enumerate(column_names, start=1):
    worksheet.cell(row=1, column=col, value=column_name)

# Iterate over the dictionaries in the list
for row, result_dict in enumerate(Results, start=2):
    # Iterate over the column names
    for col, column_name in enumerate(column_names, start=1):
        # Write the value to the appropriate cell
        value = result_dict.get(column_name, '')
        worksheet.cell(row=row, column=col, value=value)

# Save the workbook
excel_filename = 'Results-openpyxl.xlsx'
workbook.save(excel_filename)
print(f"File '{excel_filename}' created")

